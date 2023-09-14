import sys
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from datetime import datetime
from PyQt6 import QtCore
from PyQt6.QtCore import Qt
from PyQt6.QtGui import QFont, QIcon
from PyQt6.QtWidgets import QApplication, QMainWindow, QWidget, QFileDialog, QMessageBox
from tikToolsUI import Ui_MainWindow


class TableModel(QtCore.QAbstractTableModel):

    def __init__(self, data):
        super(TableModel, self).__init__()
        self._data = data

    def data(self, index, role):
        if role == Qt.ItemDataRole.DisplayRole:
            value = self._data.iloc[index.row(), index.column()]
            return str(value)

    def rowCount(self, index):
        return self._data.shape[0]

    def columnCount(self, index):
        return self._data.shape[1]

    def headerData(self, section, orientation, role):
        # section is the index of the column/row.
        if role == Qt.ItemDataRole.DisplayRole:
            if orientation == Qt.Orientation.Horizontal:
                return str(self._data.columns[section])

            if orientation == Qt.Orientation.Vertical:
                return str(self._data.index[section])



# 自定义一个窗口类
class TikTools(Ui_MainWindow, QMainWindow):
    def __init__(self):
        super().__init__()   # 调用父类的构造方法
        icon = QIcon("icon.png")
        self.setWindowIcon(icon)
        self.setupUi(self)   # 初始化构造界面
        self.btn_select_file.clicked.connect(self.open_excel_file)    # 事件绑定
        self.btn_gen_ddl.clicked.connect(self.gen_ddl_content)        # 事件绑定
        self.menu_about.triggered.connect(self.about_info)

        self.btn_gen_select.clicked.connect(self.gen_select_content)

    def open_excel_file(self):
        home_dir =  self.txt_file_name.text()
        self.fname = QFileDialog.getOpenFileName(self, 'Open file', home_dir,filter="Excel Files (*.xlsx)")

        if self.fname[0]:
            self.txt_file_name.setText(self.fname[0])

            df = pd.read_excel(self.fname[0],sheet_name='目录',usecols=[0,3,4,5],nrows=21)
            df.dropna(subset=['表名'],inplace=True)
            self.model = TableModel(df)
            self.table_file_content.setModel(self.model)

            header = self.table_file_content.horizontalHeader()
            font = QFont()
            font.setBold(True)

            header.setFont(font)

    def gen_ddl_content(self):
        excel_file = self.fname[0]
        sql_content = ""
        wb = load_workbook(excel_file, data_only=True)
        # table_name = ["ods_e3_size"]

        # 获取所有的工作表名
        sheetnames = wb.sheetnames

        # 遍历所有sheet
        for i, sheetname in enumerate(sheetnames):
            drop_table = "drop table if exists "
            create_table = "create table "
            if sheetname != '目录':
                ws = wb[sheetname]
                table_name = ws["B3"].value + "." + ws["B1"].value
                table_name_cn = ws["B2"].value
                drop_table = drop_table + table_name + ";\n"
                create_table = create_table + table_name + "( \n   "
                sql_comment = f"""/*=================================================*/\n/* 第 {i} 张表: {table_name}                        */\n/*=================================================*/\n"""
                # 从第7行开始，遍历每一行
                for j, row in enumerate(ws.iter_rows(min_row=7, values_only=True), start=7):
                    row_content = f"""{row[1]}  {row[3]}  comment '{row[2]}'"""
                    if j == 7:
                        row_content = row_content
                    else:
                        row_content = "," + row_content
                    create_table = f"""{create_table}  {row_content} \n   """

                sql_content = sql_content + sql_comment + drop_table + create_table.rstrip() + f"""\n) comment '{table_name_cn}';\n\n"""

        sql_content = f"""-- 总共生成 ddl 脚本数是：{i} \n--   生成 ddl 脚本的时间：{str(datetime.today())[:19]}\n\n""" + sql_content
        self.txt_ddl_content.setText(sql_content)

        QMessageBox.information(self, "成功提示！", f"""总共生成 ddl 脚本数是：{i}""")

    def about_info(self):
        QMessageBox.information(self, "关于", f"""Author：tik.xie\nVersion：v1.1""")

    def gen_select_content(self):
        df = pd.read_clipboard(header=0)

        # 为列数据添加后缀
        def add_suffix(s):
            column_name = s.name
            if s.dtype == "object":  # 某列为字符串类型
                return s.apply(lambda x: f"'{x}' as {column_name},")
            else:
                return s.apply(lambda x: f"{x} as {column_name},")

        df = df.apply(add_suffix)

        sql = "select "
        # 所有列都拼接成一个字符串列
        for col in df.columns:
            sql = sql + df[col]

        sql = sql.str[:-1] + " union all "
        sql.iloc[-1] = sql.iloc[-1].replace(" union all ", "")  # 处理最后一行

        sql.to_clipboard(index=False, header=False)  # 输出到剪切板
        self.txt_gen_select.setText(sql.to_string(index=False))


# 主程序入口
if __name__ == '__main__':
    app = QApplication(sys.argv)

    window = TikTools()
    window.show()

    sys.exit(app.exec())