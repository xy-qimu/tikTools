"""
 @功能描述： 根据excel数据字典文件，批量生成sql建表语句
 @创建人:    tik.xie
 @创建日期： 2023-09-06
"""

from openpyxl import load_workbook
from datetime import datetime


def gen_ddl(fname):
    wb = load_workbook(fname, data_only=True)
    sql_content = ""

    # 获取所有的工作表名
    sheetnames = wb.sheetnames

    # 遍历所有sheet
    for i, sheetname in enumerate(sheetnames):
        drop_table = "drop table if exists "
        create_table = "create table "
        if sheetname != '目录':
            ws = wb[sheetname]
            table_name = ws["B3"].value + "." + ws["B1"].value
            table_name_cn = ws["B2"].value
            drop_table = drop_table + table_name + ";\n"
            create_table = create_table + table_name + "( \n "
            comment_table = f"""comment on table {table_name} is '{table_name_cn}'; \n """
            sql_comment = f"""/*=================================================*/\n/* 第 {i} 张表: {table_name}                        */\n/*=================================================*/\n"""
            # 从第7行开始，遍历每一行
            for j, row in enumerate(ws.iter_rows(min_row=7, values_only=True), start=7):
                row_content = f"""{row[1]}  {row[3]} """
                comment_content = f"""comment on column {table_name}.{row[1]} is '{row[2]}'; \n"""
                if j == 7:
                    row_content = row_content
                else:
                    row_content = "," + row_content
                create_table = f"""{create_table}  {row_content} \n """
                comment_table = f"""{comment_table}  {comment_content} """
            sql_content = sql_content + sql_comment + drop_table + create_table.rstrip() + f"""\n) ;\n\n""" + comment_table

    sql_content = f"""-- 总共生成 ddl 脚本数是：{i} \n--   生成 ddl 脚本的时间：{str(datetime.today())[:19]}\n\n""" + sql_content

    return sql_content