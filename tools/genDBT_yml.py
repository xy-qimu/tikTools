"""
 @功能描述： 根据excel数据字典文件，批量生成dbt schema信息
 @创建人:    tik.xie
 @创建日期： 2024-04-06
"""

from openpyxl import load_workbook
from datetime import datetime


def gen_ddl(fname):
    wb = load_workbook(fname, data_only=True)
    sql_content = "version: 2\n\nmodels:\n"

    # 获取所有的工作表名
    sheetnames = wb.sheetnames

    # 遍历所有sheet
    for i, sheetname in enumerate(sheetnames):

        if sheetname != '目录':
            ws = wb[sheetname]
            table_name =  ws["B1"].value
            table_name_cn = ws["B2"].value
            sql_content = f"""{sql_content}  \n  - name: {table_name}\n    description: {table_name_cn}\n    columns:\n"""
            # 从第7行开始，遍历每一行
            for j, row in enumerate(ws.iter_rows(min_row=7, values_only=True), start=7):
                sql_content = f"""{sql_content}      - name: {row[1]}\n        description: {row[2]}\n"""

    mes_info = f"""-- 共生成 ：{i} 张表\n-- 生成时间：{str(datetime.today())[:19]}\n\n"""
    return mes_info, sql_content
